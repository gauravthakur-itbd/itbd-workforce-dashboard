import pandas as pd
import openpyxl

# Check the utilization file structure
utilization_file = 'Doc/MDE _ UTILIZATION 10-2.xlsx'

print("=" * 80)
print("ANALYZING UTILIZATION FILE STRUCTURE")
print("=" * 80)

# Load workbook to see sheets
wb = openpyxl.load_workbook(utilization_file)
print(f"\n📊 Available sheets: {wb.sheetnames}\n")

# Read each sheet
for sheet_name in wb.sheetnames:
    print(f"\n{'=' * 80}")
    print(f"SHEET: {sheet_name}")
    print(f"{'=' * 80}")
    
    try:
        df = pd.read_excel(utilization_file, sheet_name=sheet_name)
        print(f"Shape: {df.shape}")
        print(f"Columns: {list(df.columns)}\n")
        
        if not df.empty:
            print("First 10 rows:")
            print(df.head(10).to_string())
            print("\n")
    except Exception as e:
        print(f"Error reading sheet: {e}\n")

wb.close()
